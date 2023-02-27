import pandas as pd
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

from app.shared import constants
from app.utils import get_project_root


def run_service(file_name):
    """File structure example can be seen in sales_december.xlsx"""

    # read Excel
    excel_file = pd.read_excel(get_project_root() + '\\assets\\excel_automation\\' + file_name)

    # make pivot table
    report_table = excel_file.pivot_table(
        index='Gender',
        columns='Product line',
        values='Total',
        aggfunc='sum').round(0)

    # splitting the month and extension from the file name
    month_and_extension = file_name.split('_')[1]
    month_name = month_and_extension.split('.')[0]

    # prepare new Excel file for report_december.xlsx
    file_path = get_project_root() + '\\products\\' + f'report_{month_and_extension}'
    report_table.to_excel(file_path, sheet_name='Report', startrow=4)
    wb = load_workbook(file_path)
    sheet = wb['Report']

    # keep size of original
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    barchart = BarChart()
    data = Reference(sheet,
                     min_col=min_column + 1,
                     max_col=max_column,
                     min_row=min_row,
                     max_row=max_row)  # not the headers

    categories = Reference(sheet,
                           min_col=min_column,
                           max_col=min_column,
                           min_row=min_row + 1,
                           max_row=max_row)  # headers

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    sheet.add_chart(barchart, "B12")  # location chart
    barchart.title = 'Sales by Product line'
    barchart.style = 2

    # applying formulas
    # first create alphabet list as references for cells
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column]  # A,B,C columns..

    for i in excel_alphabet:
        if i != 'A':
            sheet[f'{i}{max_row + 1}'] = f'=SUM({i}{min_row + 1}:{i}{max_row})'
            sheet[f'{i}{max_row + 1}'].style = 'Currency'

    sheet[f'{excel_alphabet[0]}{max_row + 1}'] = 'Total'  # Add sum to last A cell

    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month_name.title()
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)

    wb.save(file_path)
